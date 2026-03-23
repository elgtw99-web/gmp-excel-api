from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import load_workbook
from copy import copy
import json, shutil, os, tempfile

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=False)

BASE = os.path.dirname(__file__)
TEMPLATE_PATH = os.path.join(BASE, '配方表_空白_.xlsx')
DB_PATH       = os.path.join(BASE, 'full_supplier_db.json')

with open(DB_PATH, 'r', encoding='utf-8') as f:
    _DB = json.load(f)

ING_DB  = _DB.get('ingredients', {})
MAT_DB  = _DB.get('materials', {})
PACK_DB = _DB.get('packaging', {})

# ── 查詢函數 ─────────────────────────────────────────
def _fuzzy(db, key):
    key = key.strip().upper()
    if key in db: return db[key]
    for k, v in db.items():
        if key in k or k in key: return v
    return {}

def lookup_ingredient(inci):  return _fuzzy(ING_DB,  inci)
def lookup_material(name):    return _fuzzy(MAT_DB,   name)
def lookup_packaging(name):   return _fuzzy(PACK_DB,  name)

# ── 相別推斷 ──────────────────────────────────────────
PHASE_KW = {
    'A': ['water','aqua','glycerin','propanediol','butylene glycol','pentylene glycol',
          'sodium hyaluronate','hyaluronate','niacinamide','allantoin','betaine',
          'xanthan gum','carbomer','hydroxyethylcellulose','algin','cellulose gum',
          'tranexamic acid','panthenol','dipotassium glycyrrhizinate','diglycerin',
          'sodium chloride','inositol','sodium lactate','bifida','collagen',
          'extract','ferment','peptide','oligopeptide','ascorbyl'],
    'B': ['oil','butter','wax','silicone','dimethicone','cetyl','stearyl',
          'palmitate','oleate','myristate','caprylic','squalane','jojoba',
          'tocopherol','emulsifier','peg-','polysorbate','cetearyl',
          'isononyl','ethylhexyl palmitate','hydrogenated','beeswax','lanolin',
          'cyclopentasiloxane','dimethiconol','aminomethyl propanol'],
    'C': ['phenoxyethanol','chlorphenesin','1,2-hexanediol','ethylhexylglycerin',
          'methylparaben','fragrance','parfum','hydroxyacetophenone',
          'caprylyl glycol','sodium benzoate','potassium sorbate','dehydroacetic',
          'benzyl alcohol','triclosan']
}

def infer_phase(inci, given):
    g = (given or '').strip().upper()
    if g in ('A','B','C','D'): return g
    lo = inci.lower()
    for ph, kws in PHASE_KW.items():
        for kw in kws:
            if kw in lo: return ph
    return 'A'

def copy_row_style(ws, src_row, dst_row, max_col=12):
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if src.has_style:
            dst.font          = copy(src.font)
            dst.border        = copy(src.border)
            dst.fill          = copy(src.fill)
            dst.alignment     = copy(src.alignment)
            dst.number_format = src.number_format

# ── 路由 ─────────────────────────────────────────────
@app.route('/health', methods=['GET'])
def health():
    return jsonify({
        'status': 'ok',
        'template': os.path.exists(TEMPLATE_PATH),
        'db': {'ingredients': len(ING_DB), 'materials': len(MAT_DB), 'packaging': len(PACK_DB)}
    })

@app.route('/generate-excel', methods=['POST', 'OPTIONS'])
def generate_excel():
    if request.method == 'OPTIONS':
        r = app.make_default_options_response()
        r.headers['Access-Control-Allow-Origin']  = '*'
        r.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        r.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    data = request.get_json()
    if not data: return jsonify({'error': '無效請求'}), 400

    product_name  = data.get('product_name', '新產品')
    batch_size_g  = float(data.get('batch_size_g', 62500))
    batch_no      = data.get('batch_no', '')
    date_str      = data.get('date', '')
    ingredients   = data.get('ingredients', [])   # 原料
    materials     = data.get('materials', [])      # 面膜布等物料
    packagings    = data.get('packagings', [])     # 包材
    process_steps = data.get('process_steps', [])

    if not ingredients:
        return jsonify({'error': '沒有原料資料'}), 400

    output_path = None
    try:
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.close()
        output_path = tmp.name

        shutil.copy(TEMPLATE_PATH, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        ws.title = product_name[:31]

        # 表頭
        ws['F2'] = product_name
        ws['F3'] = batch_size_g
        ws['J2'] = batch_no
        ws['J3'] = date_str

        # ── 原料分組（確保所有成分都保留）────────────
        grouped = {'A': [], 'B': [], 'C': [], 'D': []}
        for ing in ingredients:
            inci = ing.get('inci', '') or ing.get('name', '')
            ph   = infer_phase(inci, ing.get('phase', ''))
            ing['_inci'] = inci
            grouped[ph].append(ing)

        # 模板預設：A=2列(6-7), B=4列(8-11), C=3列(12-14), 合計=15
        A_START, A_DEF = 6, 2
        B_START, B_DEF = 8, 4
        C_START, C_DEF = 12, 3
        TOTAL_ROW = 15

        need_A = max(len(grouped['A']), 1)
        need_B = max(len(grouped['B']), 1)
        need_C = max(len(grouped['C']) + len(grouped['D']), 1)

        eA = max(need_A - A_DEF, 0)
        eB = max(need_B - B_DEF, 0)
        eC = max(need_C - C_DEF, 0)

        # 從後往前插入
        if eC > 0:
            ins = C_START + C_DEF
            ws.insert_rows(ins, eC)
            for r in range(ins, ins + eC): copy_row_style(ws, ins-1, r)
        if eB > 0:
            ins = B_START + B_DEF
            ws.insert_rows(ins, eB)
            for r in range(ins, ins + eB): copy_row_style(ws, ins-1, r)
        if eA > 0:
            ins = A_START + A_DEF
            ws.insert_rows(ins, eA)
            for r in range(ins, ins + eA): copy_row_style(ws, ins-1, r)

        aA = A_START
        aB = B_START + eA
        aC = C_START + eA + eB
        aT = TOTAL_ROW + eA + eB + eC
        aPK = aT + 1

        # 清空原料區
        for r in range(aA, aT):
            for c in range(1, 13):
                ws.cell(row=r, column=c).value = None

        def fill_group(label, ings, start):
            """填入一組原料，廠商查不到時欄位留空，成分絕對不丟"""
            for i, ing in enumerate(ings):
                row  = start + i
                inci = ing.get('_inci', '') or ing.get('inci', '') or ing.get('name', '')
                name = ing.get('name', '') or inci
                pct  = ing.get('percentage', None)
                sup  = lookup_ingredient(inci)

                # 有資料填資料，沒資料留空——絕不跳過這列
                ws.cell(row=row, column=1).value  = label if i == 0 else None
                ws.cell(row=row, column=2).value  = ing.get('company','')      or sup.get('company','')      or None
                ws.cell(row=row, column=3).value  = ing.get('supplierCode','') or sup.get('supplierCode','') or None
                ws.cell(row=row, column=4).value  = ing.get('productCode','')  or sup.get('productCode','')  or None
                ws.cell(row=row, column=5).value  = ing.get('batchNo','') or None
                ws.cell(row=row, column=6).value  = name if name else inci     # 至少填 INCI
                ws.cell(row=row, column=7).value  = inci or None
                if pct is not None:
                    ws.cell(row=row, column=8).value = pct
                ws.cell(row=row, column=9).value  = f'=SUM(H{row}*$F$3)'
                ws.cell(row=row, column=10).value = 'Gm'

        fill_group('A', grouped['A'], aA)
        fill_group('B', grouped['B'], aB)
        fill_group('C', grouped['C'] + grouped['D'], aC)

        # 合計
        ws.cell(row=aT, column=9).value  = f'=SUM(I{aA}:I{aT-1})'
        ws.cell(row=aT, column=10).value = 'Gm'

        # ── 包材部分 ──────────────────────────────────
        ws.cell(row=aPK, column=2).value = '包材部分'

        pack_all = list(materials) + list(packagings)
        if pack_all:
            # 找包材區起始列（aPK+1 開始填）
            for j, pk in enumerate(pack_all):
                r = aPK + 1 + j
                # 若列不存在，複製上一列樣式
                if r > ws.max_row:
                    copy_row_style(ws, aPK, r)
                name = pk.get('name', '') or pk.get('item', '')
                sup  = lookup_material(name) or lookup_packaging(name)
                ws.cell(row=r, column=2).value = pk.get('company','')      or sup.get('company','')      or None
                ws.cell(row=r, column=3).value = pk.get('supplierCode','') or sup.get('supplierCode','') or None
                ws.cell(row=r, column=4).value = pk.get('productCode','')  or sup.get('productCode','')  or None
                ws.cell(row=r, column=5).value = pk.get('batchNo','') or None
                ws.cell(row=r, column=6).value = name or None
                ws.cell(row=r, column=9).value = pk.get('quantity','') or None

        # ── 流程說明 ──────────────────────────────────
        flow_row = None
        for r in range(aPK, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            if v and '流' in str(v):
                flow_row = r
                break
        if flow_row is None:
            flow_row = aPK + len(pack_all) + 2

        if process_steps:
            steps = process_steps[:6]
        else:
            steps = []
            if grouped['A']: steps.append('A項依序加入攪拌均勻')
            if grouped['B']: steps.append('B項加入A項中攪拌均勻呈均質')
            if grouped['C']: steps.append('C項依序加入(A+B)中攪拌均勻')
            if grouped['D']: steps.append('D項依序加入混合均勻')

        ws.cell(row=flow_row, column=2).value = '流    程:'
        for j, step in enumerate(steps):
            ws.cell(row=flow_row + j, column=4).value = step

        wb.save(output_path)

        safe = product_name.replace('/', '_').replace('\\', '_')
        resp = send_file(
            output_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{safe}_秤料單.xlsx'
        )
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500
    finally:
        try:
            if output_path and os.path.exists(output_path):
                os.unlink(output_path)
        except:
            pass

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
